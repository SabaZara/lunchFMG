"""Report queries and file builders (CSV / XLSX). All file content is Georgian.

People are identified by card_id (names hidden for now). Two attendance shapes:
  * single day  → each active card marked "ჭამა" / "არ უჭამია" + a summary count.
  * multi-day   → days-attended out of days-in-range + a status.
"""
from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlmodel import Session, func, select

from .config import get_settings
from .models import Person, Scan
from .timeutil import local_date_for, local_time_str, utc_now

# Georgian labels reused across reports/exports.
L_CARD_ID = "ბარათის ID"
L_STATUS = "სტატუსი"
L_ATE = "ჭამა"
L_NOT_ATE = "არ უჭამია"
L_TIME = "დრო"
L_DATE = "თარიღი"
L_COUNT = "რაოდენობა"
L_DAYS_ATTENDED = "დასწრების დღეები"
L_DAYS_IN_RANGE = "დღეები პერიოდში"
L_TOTAL_ATE = "სულ ნაჭამი"
L_TOTAL_ACTIVE = "აქტიური ბარათები"
L_PERIOD = "პერიოდი"


# ------------------------------ summaries ---------------------------------- #
def today_summary(session: Session) -> dict:
    tz = get_settings().tz
    today = local_date_for(utc_now(), tz)
    ate = session.exec(
        select(func.count()).select_from(Scan).where(Scan.local_date == today)
    ).one()
    active = session.exec(
        select(func.count()).select_from(Person).where(Person.active == True)  # noqa: E712
    ).one()
    return {
        "date": today.isoformat(),
        "ate": int(ate),
        "active": int(active),
        "remaining": max(int(active) - int(ate), 0),
    }


def daily_counts(session: Session, frm: date, to: date) -> list[dict]:
    rows = session.exec(
        select(Scan.local_date, func.count())
        .where(Scan.local_date >= frm, Scan.local_date <= to)
        .group_by(Scan.local_date)
        .order_by(Scan.local_date)
    ).all()
    by_date = {d: int(c) for d, c in rows}
    out: list[dict] = []
    cur = frm
    while cur <= to:
        out.append({"date": cur.isoformat(), "count": by_date.get(cur, 0)})
        cur += timedelta(days=1)
    return out


@dataclass
class DayRow:
    card_id: str
    time: str


def day_detail(session: Session, day: date) -> list[DayRow]:
    """Who ate on a given local day, by card_id + local time."""
    tz = get_settings().tz
    scans = session.exec(
        select(Scan).where(Scan.local_date == day).order_by(Scan.scanned_at)
    ).all()
    return [DayRow(card_id=s.card_id, time=local_time_str(s.scanned_at, tz)) for s in scans]


def detail_rows(session: Session, frm: date, to: date) -> list[dict]:
    """Flat detail rows for any range: date, card_id, time."""
    tz = get_settings().tz
    scans = session.exec(
        select(Scan)
        .where(Scan.local_date >= frm, Scan.local_date <= to)
        .order_by(Scan.local_date, Scan.scanned_at)
    ).all()
    return [
        {
            "date": s.local_date.isoformat(),
            "card_id": s.card_id,
            "time": local_time_str(s.scanned_at, tz),
        }
        for s in scans
    ]


# --------------------------- attendance matrix ----------------------------- #
def _days_in_range(frm: date, to: date) -> int:
    return (to - frm).days + 1


def attendance(session: Session, frm: date, to: date) -> dict:
    """Compute attendance per ACTIVE card over [frm, to]."""
    active_people = session.exec(
        select(Person).where(Person.active == True).order_by(Person.card_id)  # noqa: E712
    ).all()

    # person_id -> set of local_dates attended within range
    scans = session.exec(
        select(Scan.person_id, Scan.local_date).where(
            Scan.local_date >= frm, Scan.local_date <= to
        )
    ).all()
    attended: dict[int, set] = {}
    for pid, d in scans:
        attended.setdefault(pid, set()).add(d)

    days = _days_in_range(frm, to)
    single = days == 1
    rows = []
    total_ate = 0
    for p in active_people:
        n = len(attended.get(p.id, set()))
        if n > 0:
            total_ate += 1 if single else 0
        rows.append(
            {
                "card_id": p.card_id,
                "days_attended": n,
                "attended": n > 0,
            }
        )
    if not single:
        total_ate = sum(1 for r in rows if r["days_attended"] > 0)

    return {
        "from": frm.isoformat(),
        "to": to.isoformat(),
        "days": days,
        "single_day": single,
        "rows": rows,
        "total_active": len(active_people),
        "total_ate": total_ate,
    }


# ------------------------------ CSV builders ------------------------------- #
def detail_csv(rows: list[dict]) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([L_DATE, L_CARD_ID, L_TIME])
    for r in rows:
        w.writerow([r["date"], r["card_id"], r["time"]])
    # utf-8-sig so Excel opens Georgian correctly.
    return buf.getvalue().encode("utf-8-sig")


def attendance_csv(data: dict) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    if data["single_day"]:
        w.writerow([L_CARD_ID, L_STATUS])
        for r in data["rows"]:
            w.writerow([r["card_id"], L_ATE if r["attended"] else L_NOT_ATE])
        w.writerow([])
        w.writerow([L_TOTAL_ATE, data["total_ate"]])
        w.writerow([L_TOTAL_ACTIVE, data["total_active"]])
    else:
        w.writerow([L_CARD_ID, L_DAYS_ATTENDED, L_DAYS_IN_RANGE, L_STATUS])
        for r in data["rows"]:
            status = L_ATE if r["days_attended"] > 0 else L_NOT_ATE
            w.writerow([r["card_id"], r["days_attended"], data["days"], status])
        w.writerow([])
        w.writerow([L_TOTAL_ATE, data["total_ate"]])
        w.writerow([L_TOTAL_ACTIVE, data["total_active"]])
    return buf.getvalue().encode("utf-8-sig")


# ------------------------------ XLSX builders ------------------------------ #
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_ATE_FILL = PatternFill("solid", fgColor="C6EFCE")
_NOT_FILL = PatternFill("solid", fgColor="FFC7CE")


def _style_header(ws, ncols: int) -> None:  # noqa: ANN001
    for col in range(1, ncols + 1):
        c = ws.cell(row=1, column=col)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = Alignment(horizontal="center")


def _autofit(ws, widths: list[int]) -> None:  # noqa: ANN001
    from openpyxl.utils import get_column_letter

    for i, wdt in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = wdt


def _force_text(ws) -> None:  # noqa: ANN001
    """Force the card-id column (col 1) to text so leading zeros are preserved."""
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
        for cell in row:
            cell.number_format = "@"


def detail_xlsx(rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "დეტალები"
    ws.append([L_DATE, L_CARD_ID, L_TIME])
    for r in rows:
        ws.append([r["date"], str(r["card_id"]), r["time"]])
    _style_header(ws, 3)
    _autofit(ws, [14, 22, 12])
    # card_id is column 2 here; force it to text.
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = "@"
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def attendance_xlsx(data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "დასწრება"

    if data["single_day"]:
        ws.append([L_CARD_ID, L_STATUS])
        _style_header(ws, 2)
        for r in data["rows"]:
            status = L_ATE if r["attended"] else L_NOT_ATE
            ws.append([str(r["card_id"]), status])
            ws.cell(row=ws.max_row, column=2).fill = (
                _ATE_FILL if r["attended"] else _NOT_FILL
            )
        _autofit(ws, [22, 16])
    else:
        ws.append([L_CARD_ID, L_DAYS_ATTENDED, L_DAYS_IN_RANGE, L_STATUS])
        _style_header(ws, 4)
        for r in data["rows"]:
            attended = r["days_attended"] > 0
            status = L_ATE if attended else L_NOT_ATE
            ws.append([str(r["card_id"]), r["days_attended"], data["days"], status])
            ws.cell(row=ws.max_row, column=4).fill = (
                _ATE_FILL if attended else _NOT_FILL
            )
        _autofit(ws, [22, 18, 16, 16])

    _force_text(ws)

    # Summary block below a blank row.
    ws.append([])
    ws.append([L_PERIOD, f'{data["from"]} — {data["to"]}'])
    ws.append([L_TOTAL_ATE, data["total_ate"]])
    ws.append([L_TOTAL_ACTIVE, data["total_active"]])

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
