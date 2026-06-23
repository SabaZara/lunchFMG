"""Bulk import of card IDs from .xlsx or .csv.

Format: ONE CARD ID PER LINE in the first column. An optional "card_id" header
is tolerated. Each row becomes an ACTIVE person with full_name = "----".

CRITICAL: cells are read as TEXT so leading zeros survive (Excel loves to turn
"0573856032" into the number 573856032). We coerce ints to str without losing
zeros where openpyxl already typed them, and we strip whitespace / BOM.

Returns a structured report: added count, and per-row failures/duplicates by
1-based row number (matching what the operator sees in Excel).
"""
from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field

from openpyxl import load_workbook
from sqlmodel import Session, select

from .models import NAME_PLACEHOLDER, Person

_HEADER_TOKENS = {"card_id", "cardid", "card", "ბარათი", "id"}


@dataclass
class ImportReport:
    added: int = 0
    duplicates: list[tuple[int, str]] = field(default_factory=list)   # (row, card_id)
    invalid: list[tuple[int, str]] = field(default_factory=list)      # (row, reason)
    total_rows: int = 0

    def as_dict(self) -> dict:
        return {
            "added": self.added,
            "duplicate_count": len(self.duplicates),
            "invalid_count": len(self.invalid),
            "total_rows": self.total_rows,
            "duplicates": [{"row": r, "card_id": c} for r, c in self.duplicates],
            "invalid": [{"row": r, "reason": reason} for r, reason in self.invalid],
        }


def _clean_cell(value) -> str:  # noqa: ANN001
    """Coerce a cell to a clean card-id string, preserving leading zeros."""
    if value is None:
        return ""
    if isinstance(value, float):
        # Excel may give 1001.0 — keep an integer-looking value without ".0".
        if value.is_integer():
            return str(int(value))
        return str(value)
    if isinstance(value, int):
        return str(value)
    text = str(value)
    # Strip BOM and surrounding whitespace; keep internal characters/zeros.
    return text.lstrip("﻿").strip()


def _rows_from_xlsx(data: bytes) -> list[str]:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    out: list[str] = []
    for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
        out.append(_clean_cell(row[0] if row else None))
    wb.close()
    return out


def _rows_from_csv(data: bytes) -> list[str]:
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    out: list[str] = []
    for row in reader:
        out.append(_clean_cell(row[0]) if row else "")
    return out


def parse_rows(filename: str, data: bytes) -> list[str]:
    name = (filename or "").lower()
    if name.endswith(".csv"):
        return _rows_from_csv(data)
    # default to xlsx
    return _rows_from_xlsx(data)


def import_cards(session: Session, filename: str, data: bytes) -> ImportReport:
    rows = parse_rows(filename, data)
    report = ImportReport()

    # Track ids seen within THIS file so a dup inside the file is reported too.
    seen_in_file: set[str] = set()

    for idx, raw in enumerate(rows, start=1):
        card_id = raw
        if card_id == "":
            # Skip blank lines silently (not counted as a real row).
            continue
        report.total_rows += 1

        # Tolerate a single header line.
        if idx == 1 and card_id.lower() in _HEADER_TOKENS:
            report.total_rows -= 1
            continue

        if card_id in seen_in_file:
            report.duplicates.append((idx, card_id))
            continue
        seen_in_file.add(card_id)

        existing = session.exec(
            select(Person).where(Person.card_id == card_id)
        ).first()
        if existing is not None:
            report.duplicates.append((idx, card_id))
            continue

        session.add(
            Person(card_id=card_id, full_name=NAME_PLACEHOLDER, active=True)
        )
        report.added += 1

    if report.added:
        session.commit()
    return report
