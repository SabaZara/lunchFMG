"""Test the kiosk scan flow WITHOUT a physical card reader.

Reads real card IDs from tests/data/real_cards.xlsx (the same file you imported)
and POSTs them to the LIVE local app at http://127.0.0.1:<PORT>/api/scan — the
exact code path a USB reader triggers when it "types" a card id + Enter.

Run ON THE KIOSK (the app must already be running via start.bat):

    .venv\\Scripts\\python.exe scripts\\test_scan.py            (scan first 5 real cards)
    .venv\\Scripts\\python.exe scripts\\test_scan.py --count 10  (first 10)
    .venv\\Scripts\\python.exe scripts\\test_scan.py 0573856032  (scan specific id(s))
    .venv\\Scripts\\python.exe scripts\\test_scan.py --all       (every real card — careful, marks them eaten)
    .venv\\Scripts\\python.exe scripts\\test_scan.py --dry-run    (READ-ONLY: check status, record nothing)

Nothing is imported or changed in the card list; a real scan DOES record a meal
for that card for today (by design). Use --dry-run to verify card data WITHOUT
consuming anyone's daily slot (it reads the DB directly and the app need not be
running). Use a small --count for a real smoke test.
"""
from __future__ import annotations

import argparse
import json
import sys
import urllib.error
import urllib.request
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from app.config import get_settings  # noqa: E402

REAL_CARDS = ROOT / "tests" / "data" / "real_cards.xlsx"


def _load_card_ids(limit: int | None) -> list[str]:
    from openpyxl import load_workbook

    if not REAL_CARDS.exists():
        print(f"[test_scan] card file not found: {REAL_CARDS}")
        return []
    wb = load_workbook(REAL_CARDS, read_only=True, data_only=True)
    ws = wb.active
    ids: list[str] = []
    seen = set()
    for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
        v = row[0]
        if v is None:
            continue
        s = str(int(v)) if isinstance(v, float) and v.is_integer() else str(v)
        s = s.strip()
        if not s or s.lower() in {"card_id", "cardid", "card", "id", "ბარათი"}:
            continue
        if s not in seen:
            seen.add(s)
            ids.append(s)
    wb.close()
    return ids if limit is None else ids[:limit]


def _scan(base: str, card_id: str) -> dict:
    data = json.dumps({"card_id": card_id}).encode("utf-8")
    req = urllib.request.Request(
        f"{base}/api/scan",
        data=data,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=10) as resp:
        return json.loads(resp.read().decode("utf-8"))


def _dry_run(args) -> int:  # noqa: ANN001
    """Read-only check straight against the DB — records nothing."""
    from sqlmodel import Session, select

    from app.db import engine
    from app.models import Person, Scan
    from app.timeutil import local_date_for, utc_now

    cards = args.card_ids or _load_card_ids(None if args.all else args.count)
    if not cards:
        print("[test_scan] no card ids to check.")
        return 1

    today = local_date_for(utc_now(), get_settings().tz)
    print(f"[test_scan] DRY RUN (read-only) — {len(cards)} card(s), local date {today}\n")
    would_allow = would_deny = 0
    with Session(engine) as s:
        for cid in cards:
            person = s.exec(select(Person).where(Person.card_id == cid.strip())).first()
            if person is None:
                print(f"  {cid:<16} would DENY  უცნობი ბარათი (not in DB)")
                would_deny += 1
                continue
            if not person.active:
                print(f"  {cid:<16} would DENY  ბარათი გათიშულია (inactive)")
                would_deny += 1
                continue
            ate = s.exec(
                select(Scan).where(Scan.person_id == person.id, Scan.local_date == today)
            ).first()
            if ate:
                print(f"  {cid:<16} would DENY  დღეს უკვე ნაჭამია (already ate today)")
                would_deny += 1
            else:
                print(f"  {cid:<16} would ALLOW (active, not yet eaten today)")
                would_allow += 1
    print(f"\n[test_scan] dry-run: {would_allow} would-allow, {would_deny} would-deny. "
          "Nothing was recorded.")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description="Test kiosk scans without a reader.")
    parser.add_argument("card_ids", nargs="*", help="specific card id(s) to scan")
    parser.add_argument("--count", type=int, default=5, help="how many real cards to scan (default 5)")
    parser.add_argument("--all", action="store_true", help="scan every real card")
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="READ-ONLY: report each card's status from the DB without recording "
        "a meal (does not consume anyone's daily slot, app need not be running)",
    )
    args = parser.parse_args()

    if args.dry_run:
        return _dry_run(args)

    port = get_settings().port
    base = f"http://{args.host}:{port}"

    if args.card_ids:
        cards = args.card_ids
    else:
        cards = _load_card_ids(None if args.all else args.count)

    if not cards:
        print("[test_scan] no card ids to scan.")
        return 1

    # Confirm the app is up.
    try:
        with urllib.request.urlopen(f"{base}/healthz", timeout=5) as r:
            r.read()
    except (urllib.error.URLError, OSError) as exc:
        print(f"[test_scan] cannot reach the app at {base} — is start.bat running? ({exc})")
        return 1

    print(f"[test_scan] scanning {len(cards)} card(s) against {base}\n")
    allowed = denied = errors = 0
    for cid in cards:
        try:
            res = _scan(base, cid)
        except Exception as exc:  # noqa: BLE001
            print(f"  {cid:<16} ERROR {exc}")
            errors += 1
            continue
        status = res.get("status")
        if status == "ALLOWED":
            allowed += 1
            print(f"  {cid:<16} ALLOWED  ({res.get('scanned_at')})")
        else:
            denied += 1
            print(f"  {cid:<16} DENIED   {res.get('reason')}")

    print(f"\n[test_scan] done: {allowed} allowed, {denied} denied, {errors} errors.")
    print("[test_scan] tip: re-running the same cards today should show "
          "'დღეს უკვე ნაჭამია' (already eaten).")
    return 0


if __name__ == "__main__":
    sys.exit(main())
