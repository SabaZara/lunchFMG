"""Automated acceptance tests covering spec section 16 (Definition of Done).

Card IDs are typed (no physical reader), exactly as in production testing.
"""
from __future__ import annotations

import io
import threading
from datetime import date, timedelta
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from sqlmodel import Session, select

# Real (gitignored) production card file, bundled as a fixture on machines that
# have it. Tests using it skip cleanly when it is absent (e.g. clean checkout).
REAL_CARDS_FILE = Path(__file__).resolve().parent / "data" / "real_cards.xlsx"


# --------------------------- helpers --------------------------------------- #
def _login(ctx):
    r = ctx["client"].post(
        "/api/login",
        headers=ctx["headers"],
        json={"username": ctx["admin_user"], "password": ctx["admin_pass"]},
    )
    assert r.status_code == 200, r.text


def _seed_cards(ctx):
    with Session(ctx["db"].engine) as s:
        ctx["seed"].seed_sample_cards(s)


def _make_xlsx(values, header=None):
    wb = Workbook()
    ws = wb.active
    if header:
        ws.append([header])
    for v in values:
        ws.append([v])  # one card id per line, first column
    # Force text format so openpyxl doesn't coerce to number on read-back.
    for row in ws.iter_rows(min_col=1, max_col=1):
        for c in row:
            c.number_format = "@"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --------------------------- scan behaviour -------------------------------- #
def test_known_active_card_allowed_and_records(app_ctx):
    ctx = app_ctx
    _seed_cards(ctx)
    r = ctx["client"].post("/api/scan", json={"card_id": "1001"})
    j = r.json()
    assert j["status"] == "ALLOWED"
    assert j["scanned_at"]  # local time string present
    # default daily limit is 2 -> after the first meal, 1 remains
    assert j["remaining"] == 1 and j["limit"] == 2


def test_daily_limit_two_meals_then_denied(app_ctx):
    """Default limit 2: two meals allowed, the third denied (limit reached)."""
    ctx = app_ctx
    _seed_cards(ctx)
    c = ctx["client"]
    j1 = c.post("/api/scan", json={"card_id": "1001"}).json()
    assert j1["status"] == "ALLOWED" and j1["remaining"] == 1
    j2 = c.post("/api/scan", json={"card_id": "1001"}).json()
    assert j2["status"] == "ALLOWED" and j2["remaining"] == 0
    j3 = c.post("/api/scan", json={"card_id": "1001"}).json()
    assert j3["status"] == "DENIED"
    assert j3["reason"] == "დღის ლიმიტი ამოიწურა"


def test_daily_limit_one_respected(app_ctx):
    """A card set to limit 1 behaves like the old once-per-day rule."""
    ctx = app_ctx
    _login(ctx)
    H = ctx["headers"]
    c = ctx["client"]
    pid = c.post("/api/people", headers=H, json={"card_id": "ONE1", "daily_limit": 1}).json()["id"]
    assert c.post("/api/scan", json={"card_id": "ONE1"}).json()["status"] == "ALLOWED"
    j = c.post("/api/scan", json={"card_id": "ONE1"}).json()
    assert j["status"] == "DENIED" and j["reason"] == "დღის ლიმიტი ამოიწურა"


def test_unknown_card_denied(app_ctx):
    ctx = app_ctx
    _seed_cards(ctx)
    j = ctx["client"].post("/api/scan", json={"card_id": "DOES-NOT-EXIST"}).json()
    assert j["status"] == "DENIED"
    assert j["reason"] == "უცნობი ბარათი"


def test_inactive_card_denied(app_ctx):
    ctx = app_ctx
    _seed_cards(ctx)
    j = ctx["client"].post("/api/scan", json={"card_id": "9999"}).json()
    assert j["status"] == "DENIED"
    assert j["reason"] == "ბარათი გათიშულია"


def test_midnight_reset(app_ctx):
    """A card allowed 'yesterday' is allowed again 'today'.

    We simulate by inserting a scan row with yesterday's local_date directly,
    then a fresh scan today must be ALLOWED.
    """
    ctx = app_ctx
    _seed_cards(ctx)
    from app.models import Person, Scan
    from app.timeutil import local_date_for, utc_now

    with Session(ctx["db"].engine) as s:
        p = s.exec(select(Person).where(Person.card_id == "1002")).first()
        tz = ctx["settings"].tz
        today = local_date_for(utc_now(), tz)
        yesterday = today - timedelta(days=1)
        s.add(Scan(person_id=p.id, card_id="1002", scanned_at=utc_now(), local_date=yesterday))
        s.commit()

    # Today's scan should still be allowed.
    j = ctx["client"].post("/api/scan", json={"card_id": "1002"}).json()
    assert j["status"] == "ALLOWED"


def test_concurrent_taps_respect_limit(app_ctx):
    """Under concurrency the daily limit holds: exactly `limit` ALLOWED."""
    ctx = app_ctx
    _seed_cards(ctx)
    results = []
    lock = threading.Lock()

    def tap():
        r = ctx["client"].post("/api/scan", json={"card_id": "1003"})
        with lock:
            results.append(r.json()["status"])

    threads = [threading.Thread(target=tap) for _ in range(12)]
    for t in threads:
        t.start()
    for t in threads:
        t.join()

    allowed = sum(1 for s in results if s == "ALLOWED")
    # seeded cards use the default limit of 2
    assert allowed == 2, f"expected exactly 2 ALLOWED, got {allowed}: {results}"


def test_leading_zeros_preserved_through_scan(app_ctx):
    ctx = app_ctx
    _seed_cards(ctx)
    j = ctx["client"].post("/api/scan", json={"card_id": "0573856032"}).json()
    assert j["status"] == "ALLOWED"
    # Second tap still allowed (limit 2); third denied.
    j2 = ctx["client"].post("/api/scan", json={"card_id": "0573856032"}).json()
    assert j2["status"] == "ALLOWED"
    j3 = ctx["client"].post("/api/scan", json={"card_id": "0573856032"}).json()
    assert j3["status"] == "DENIED"


# --------------------------- admin / CRUD ---------------------------------- #
def test_admin_crud_and_unique_enforcement(app_ctx):
    ctx = app_ctx
    _login(ctx)
    H = ctx["headers"]
    c = ctx["client"]

    # add
    r = c.post("/api/people", headers=H, json={"card_id": "AAA111"})
    assert r.status_code == 201
    pid = r.json()["id"]
    assert r.json()["full_name"] == "----"  # placeholder default

    # duplicate rejected with Georgian error
    r = c.post("/api/people", headers=H, json={"card_id": "AAA111"})
    assert r.status_code == 409
    assert "მინიჭებულია" in r.json()["detail"]

    # edit (reassign card id)
    r = c.put(f"/api/people/{pid}", headers=H, json={"card_id": "BBB222"})
    assert r.status_code == 200 and r.json()["card_id"] == "BBB222"

    # deactivate
    r = c.put(f"/api/people/{pid}", headers=H, json={"active": False})
    assert r.json()["active"] is False

    # delete
    r = c.delete(f"/api/people/{pid}", headers=H)
    assert r.status_code == 204


def test_admin_toggle_ate_today(app_ctx):
    """Admin can mark/un-mark whether a person ate today; reflects in scan logic."""
    ctx = app_ctx
    _login(ctx)
    H = ctx["headers"]
    c = ctx["client"]

    r = c.post("/api/people", headers=H, json={"card_id": "EAT001"})
    pid = r.json()["id"]
    assert r.json()["ate_today"] is False and r.json()["ate_count"] == 0
    assert r.json()["daily_limit"] == 2

    # mark eaten -> fills up to the daily limit (2)
    r = c.post(f"/api/people/{pid}/ate", headers=H, json={"ate": True})
    assert r.status_code == 200 and r.json()["ate_count"] == 2

    # marking again is idempotent (already at the limit)
    r = c.post(f"/api/people/{pid}/ate", headers=H, json={"ate": True})
    assert r.json()["ate_count"] == 2

    # a real kiosk scan now reports limit reached (consistent with the manual mark)
    j = c.post("/api/scan", json={"card_id": "EAT001"}).json()
    assert j["status"] == "DENIED" and j["reason"] == "დღის ლიმიტი ამოიწურა"

    # un-mark -> clears all of today's meals
    r = c.post(f"/api/people/{pid}/ate", headers=H, json={"ate": False})
    assert r.status_code == 200 and r.json()["ate_count"] == 0

    # now the kiosk allows the meal again
    j = c.post("/api/scan", json={"card_id": "EAT001"}).json()
    assert j["status"] == "ALLOWED"

    # endpoint is gated (no secret -> 403)
    assert c.post(f"/api/people/{pid}/ate", json={"ate": True}).status_code == 403


def test_bulk_operations_and_export(app_ctx):
    """Bulk delete / activate / deactivate / ate / unate + CSV export."""
    ctx = app_ctx
    _login(ctx)
    H = ctx["headers"]
    c = ctx["client"]

    ids = []
    for cid in ["B01", "B02", "B03", "B04"]:
        ids.append(c.post("/api/people", headers=H, json={"card_id": cid}).json()["id"])

    # bulk deactivate two
    r = c.post("/api/people/bulk", headers=H,
               json={"action": "deactivate", "ids": ids[:2]})
    assert r.json()["affected"] == 2
    listed = {p["card_id"]: p for p in c.get("/api/people", headers=H).json()}
    assert listed["B01"]["active"] is False and listed["B03"]["active"] is True

    # bulk mark ate for three
    r = c.post("/api/people/bulk", headers=H, json={"action": "ate", "ids": ids[:3]})
    assert r.json()["affected"] == 3
    listed = {p["card_id"]: p for p in c.get("/api/people", headers=H).json()}
    assert listed["B01"]["ate_today"] and listed["B03"]["ate_today"] and not listed["B04"]["ate_today"]

    # bulk un-ate two
    r = c.post("/api/people/bulk", headers=H, json={"action": "unate", "ids": ids[:2]})
    assert r.json()["affected"] == 2

    # CSV export contains Georgian headers + a card id
    exp = c.get("/api/people/export.csv", headers=H)
    assert exp.status_code == 200
    text = exp.content.decode("utf-8-sig")
    assert "ბარათის ID" in text and "B04" in text

    # bulk delete selected
    r = c.post("/api/people/bulk", headers=H, json={"action": "delete", "ids": ids[:2]})
    assert r.json()["affected"] == 2
    remaining = {p["card_id"] for p in c.get("/api/people", headers=H).json()}
    assert "B01" not in remaining and "B03" in remaining

    # delete ALL
    r = c.post("/api/people/bulk", headers=H, json={"action": "delete", "all": True})
    assert r.json()["affected"] >= 1
    assert c.get("/api/people", headers=H).json() == []

    # unknown action rejected; endpoint gated
    assert c.post("/api/people/bulk", headers=H, json={"action": "nope"}).status_code == 422
    assert c.post("/api/people/bulk", json={"action": "delete", "all": True}).status_code == 403
    assert c.get("/api/people/export.csv").status_code == 403


def test_import_250_xlsx_leading_zeros(app_ctx):
    ctx = app_ctx
    _login(ctx)
    H = ctx["headers"]
    c = ctx["client"]

    ids = [f"{i:010d}" for i in range(1, 251)]  # 250 ids, all with leading zeros
    data = _make_xlsx(ids, header="card_id")  # tolerate header
    r = c.post(
        "/api/people/import",
        headers=H,
        files={"file": ("cards.xlsx", data,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    rep = r.json()
    assert rep["added"] == 250, rep
    assert rep["duplicate_count"] == 0

    # leading zeros preserved in DB
    listed = c.get("/api/people?q=0000000001", headers=H).json()
    assert any(p["card_id"] == "0000000001" for p in listed)

    # re-import same file -> all duplicates, none added
    r2 = c.post(
        "/api/people/import",
        headers=H,
        files={"file": ("cards.xlsx", data,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r2.json()["added"] == 0
    assert r2.json()["duplicate_count"] == 250


def test_import_reports_failures_by_row(app_ctx):
    ctx = app_ctx
    _login(ctx)
    H = ctx["headers"]
    c = ctx["client"]
    # rows: header, A, blank(skip), A(dup), B
    wb = Workbook(); ws = wb.active
    ws.append(["card_id"])
    ws.append(["A001"])
    ws.append([""])         # blank skipped
    ws.append(["A001"])     # duplicate within file -> row 4
    ws.append(["B002"])
    buf = io.BytesIO(); wb.save(buf)
    r = c.post("/api/people/import", headers=H,
               files={"file": ("c.xlsx", buf.getvalue(), "application/octet-stream")})
    rep = r.json()
    assert rep["added"] == 2
    assert rep["duplicate_count"] == 1
    assert rep["duplicates"][0]["row"] == 4


def test_csv_import_bom_tolerant(app_ctx):
    ctx = app_ctx
    _login(ctx)
    H = ctx["headers"]
    c = ctx["client"]
    csv_bytes = "﻿card_id\n0012\n0034\n".encode("utf-8")
    r = c.post("/api/people/import", headers=H,
               files={"file": ("c.csv", csv_bytes, "text/csv")})
    rep = r.json()
    assert rep["added"] == 2
    listed = c.get("/api/people?q=0012", headers=H).json()
    assert any(p["card_id"] == "0012" for p in listed)


def test_import_real_world_excel_hazards(app_ctx):
    """A real ~250-card upload mixes string cells, NUMBER-typed cells (Excel
    auto-types them), floats, and stray whitespace. Card IDs must survive as
    text with leading zeros intact, and a card scanned afterward must match.
    """
    ctx = app_ctx
    _login(ctx)
    H = ctx["headers"]
    c = ctx["client"]

    wb = Workbook()
    ws = wb.active
    ws.append(["card_id"])          # header tolerated
    ws.append([1001])               # integer-typed cell -> "1001"
    ws.append([1002.0])             # float-typed cell    -> "1002"
    ws.append(["  0573856032  "])   # whitespace + leading zeros -> "0573856032"
    ws.append([" 0044 "])           # whitespace + leading zeros -> "0044"
    ws.append([""])                 # blank -> skipped
    buf = io.BytesIO()
    wb.save(buf)

    r = c.post("/api/people/import", headers=H,
               files={"file": ("cards.xlsx", buf.getvalue(),
                               "application/octet-stream")})
    rep = r.json()
    assert rep["added"] == 4, rep

    listed = {p["card_id"] for p in c.get("/api/people", headers=H).json()}
    assert {"1001", "1002", "0573856032", "0044"} <= listed

    # The whitespace-padded, leading-zero card scans correctly at the kiosk.
    j = c.post("/api/scan", json={"card_id": "0573856032"}).json()
    assert j["status"] == "ALLOWED"


@pytest.mark.skipif(not REAL_CARDS_FILE.exists(),
                    reason="real_cards.xlsx fixture not present")
def test_import_real_card_file(app_ctx):
    """Import the actual production .xlsx and verify the end-to-end contract:
    every card lands as text (leading zeros intact), no duplicates, and a known
    leading-zero card both imports and scans. Re-import yields all duplicates.
    """
    ctx = app_ctx
    _login(ctx)
    H = ctx["headers"]
    c = ctx["client"]

    data = REAL_CARDS_FILE.read_bytes()

    # What the file actually contains (computed, so the test isn't hard-coded
    # to a row count and survives the file being updated).
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    raw = [row[0] for row in ws.iter_rows(min_col=1, max_col=1, values_only=True)]
    expected = []
    seen = set()
    for v in raw:
        if v is None:
            continue
        s = str(int(v)) if isinstance(v, float) and v.is_integer() else str(v)
        s = s.strip()
        if s and s.lower() not in {"card_id", "cardid", "card", "id", "ბარათი"}:
            if s not in seen:
                seen.add(s)
                expected.append(s)
    wb.close()
    assert expected, "fixture appears empty"

    r = c.post("/api/people/import", headers=H,
               files={"file": ("real_cards.xlsx", data,
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    rep = r.json()
    assert rep["added"] == len(expected), rep
    assert rep["duplicate_count"] == 0
    assert rep["invalid_count"] == 0

    listed = {p["card_id"] for p in c.get("/api/people", headers=H).json()}
    assert set(expected) <= listed
    # Leading zeros must be preserved verbatim for any 0-prefixed card.
    zero_cards = [e for e in expected if e.startswith("0")]
    if zero_cards:
        assert zero_cards[0] in listed
        assert c.post("/api/scan", json={"card_id": zero_cards[0]}).json()["status"] == "ALLOWED"

    # Re-importing the same file adds nothing and reports all as duplicates.
    r2 = c.post("/api/people/import", headers=H,
                files={"file": ("real_cards.xlsx", data,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r2.json()["added"] == 0
    assert r2.json()["duplicate_count"] == len(expected)


# --------------------------- reports / exports ----------------------------- #
def test_today_report_counts(app_ctx):
    ctx = app_ctx
    _seed_cards(ctx)
    _login(ctx)
    ctx["client"].post("/api/scan", json={"card_id": "1001"})
    t = ctx["client"].get("/api/reports/today", headers=ctx["headers"]).json()
    assert t["ate"] == 1
    assert t["active"] >= 1
    assert t["remaining"] == t["active"] - t["ate"]


def test_reports_count_people_vs_meals(app_ctx):
    """Two meals by one person: today shows 1 person / 2 meals; the day view
    groups that card into ONE row with count 2 and both times."""
    ctx = app_ctx
    _seed_cards(ctx)
    _login(ctx)
    H = ctx["headers"]
    c = ctx["client"]
    # 1001 has default limit 2 -> eat twice
    c.post("/api/scan", json={"card_id": "1001"})
    c.post("/api/scan", json={"card_id": "1001"})

    t = c.get("/api/reports/today", headers=H).json()
    assert t["people_ate"] == 1 and t["meals"] == 2
    assert t["remaining"] == t["active"] - 1  # one person ate (not two)

    today = date.today().isoformat()
    day = c.get(f"/api/reports/day?date={today}", headers=H).json()
    assert day["people"] == 1 and day["meals"] == 2
    assert len(day["rows"]) == 1
    row = day["rows"][0]
    assert row["card_id"] == "1001" and row["count"] == 2 and len(row["times"]) == 2


def test_attendance_xlsx_is_georgian_and_by_cardid(app_ctx):
    ctx = app_ctx
    _seed_cards(ctx)
    _login(ctx)
    H = ctx["headers"]
    ctx["client"].post("/api/scan", json={"card_id": "1001"})
    today = date.today().isoformat()
    r = ctx["client"].get(f"/api/reports/attendance?from={today}&to={today}&format=xlsx", headers=H)
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert "ბარათის ID" in headers and "სტატუსი" in headers
    # status cells use Georgian "ჭამა" / "არ უჭამია"
    statuses = {ws.cell(row=i, column=2).value for i in range(2, ws.max_row + 1)}
    assert statuses & {"ჭამა", "არ უჭამია"}


def test_detail_csv_export(app_ctx):
    ctx = app_ctx
    _seed_cards(ctx)
    _login(ctx)
    H = ctx["headers"]
    ctx["client"].post("/api/scan", json={"card_id": "0573856032"})
    today = date.today().isoformat()
    r = ctx["client"].get(f"/api/reports/export?from={today}&to={today}&format=csv", headers=H)
    assert r.status_code == 200
    text = r.content.decode("utf-8-sig")
    assert "ბარათის ID" in text
    assert "0573856032" in text  # leading zeros preserved in export


# --------------------------- security / gating ----------------------------- #
def test_admin_blocked_without_tunnel_secret(app_ctx):
    ctx = app_ctx
    c = ctx["client"]
    # No secret header -> 403 on all protected surfaces.
    assert c.get("/admin").status_code == 403
    assert c.get("/reports").status_code == 403
    assert c.post("/api/login", json={"username": "x", "password": "y"}).status_code == 403
    assert c.get("/api/people").status_code == 403
    assert c.get("/api/reports/today").status_code == 403


def test_version_endpoint(app_ctx):
    ctx = app_ctx
    from app import __version__
    r = ctx["client"].get("/api/version")  # always-open, no secret needed
    assert r.status_code == 200
    assert r.json()["version"] == __version__


def test_update_endpoint_gated_and_status(app_ctx):
    """Remote update is gated; status reports version+repo; POST applies without
    actually pulling/restarting (subprocess mocked)."""
    import subprocess as _sp
    ctx = app_ctx
    c = ctx["client"]
    H = ctx["headers"]

    # gated: no secret -> 403
    assert c.get("/api/update/status").status_code == 403
    assert c.post("/api/update").status_code == 403

    _login(ctx)
    s = c.get("/api/update/status", headers=H).json()
    assert "version" in s and "repo" in s

    # Mock the apply step (success) and the detached restart spawn so the test
    # neither hits GitHub nor restarts anything.
    import app.routers.update as upd

    class _OK:
        returncode = 0
        stdout = "[update] applied 5 files\n"
        stderr = ""

    spawned = {"called": False}
    orig_run, orig_popen = upd.subprocess.run, upd.subprocess.Popen
    upd.subprocess.run = lambda *a, **k: _OK()
    upd.subprocess.Popen = lambda *a, **k: spawned.__setitem__("called", True)
    try:
        r = c.post("/api/update", headers=H)
    finally:
        upd.subprocess.run, upd.subprocess.Popen = orig_run, orig_popen

    j = r.json()
    assert r.status_code == 200 and j["ok"] and j["applied"]
    assert j["restarting"] is True and spawned["called"] is True
    assert "applied" in j["output"]


def test_scan_page_and_api_always_open(app_ctx):
    ctx = app_ctx
    c = ctx["client"]
    # No secret header, kiosk still works (offline behaviour).
    assert c.get("/").status_code == 200
    assert c.post("/api/scan", json={"card_id": "whatever"}).status_code == 200


def test_gate_denies_unknown_paths_by_default(app_ctx):
    ctx = app_ctx
    c = ctx["client"]
    # An unlisted path (no allowlist match) must require the secret, not fall
    # through to allowed. With the secret it 404s (no such route) but is NOT
    # blocked by the gate; without it, 403.
    assert c.get("/totally-unknown-path").status_code == 403
    assert c.get("/api/unknown").status_code == 403
    # Static helper assets for the test page are local-only too.
    assert c.get("/static/kiosk-test.js", headers=ctx["headers"]).status_code == 403


def test_kiosk_test_page_is_local_only(app_ctx):
    ctx = app_ctx
    c = ctx["client"]
    assert c.get("/kiosk-test").status_code == 200
    assert c.get("/kiosk-test", headers=ctx["headers"]).status_code == 403


def test_login_rate_limit(app_ctx):
    ctx = app_ctx
    H = ctx["headers"]
    c = ctx["client"]
    last = None
    for _ in range(6):
        last = c.post("/api/login", headers=H,
                      json={"username": ctx["admin_user"], "password": "wrong"})
    # After 5 failures the 6th is locked out (429).
    assert last.status_code == 429
    assert "მცდელობა" in last.json()["detail"]


def test_weak_password_refused():
    """validate_settings refuses a weak admin password."""
    from app.config import ConfigError, Settings, validate_settings

    s = Settings(
        timezone="Asia/Tbilisi",
        admin_username="admin",
        admin_password="changeme",
        secret_key="x" * 48,
        tunnel_secret="t" * 16,
        host="127.0.0.1",
        port=8000,
        db_path=__import__("pathlib").Path("x.db"),
        seed_sample_cards=False,
    )
    with pytest.raises(ConfigError):
        validate_settings(s)


def test_missing_secret_key_refused():
    from app.config import ConfigError, Settings, validate_settings

    s = Settings(
        timezone="Asia/Tbilisi",
        admin_username="admin",
        admin_password="StrongPass!123",
        secret_key="short",
        tunnel_secret="t" * 16,
        host="127.0.0.1",
        port=8000,
        db_path=__import__("pathlib").Path("x.db"),
        seed_sample_cards=False,
    )
    with pytest.raises(ConfigError):
        validate_settings(s)
